"""Generate state-wise Excel files from JustDial scraped_data.json into output_justdial folder."""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def main():
    data_file = os.path.join("output", "scraped_data.json")
    output_dir = "output_justdial"

    with open(data_file, "r", encoding="utf-8") as f:
        all_items = json.load(f)

    # Filter JustDial only
    jd_items = [item for item in all_items if item.get("source") == "JustDial"]
    print(f"Total JustDial records: {len(jd_items)}")

    # Group by state
    state_data = {}
    for item in jd_items:
        state = item.get("state", "Unknown")
        if state not in state_data:
            state_data[state] = []
        state_data[state].append(item)

    # Styles
    hf = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center")
    tb = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    df = Font(name="Calibri", size=11)

    headers = ["S.No", "Name", "Phone", "Address", "District", "City", "Pincode", "State", "Source"]
    col_widths = [8, 35, 18, 45, 20, 18, 10, 20, 12]

    os.makedirs(output_dir, exist_ok=True)

    # Individual state Excel files
    for state in sorted(state_data.keys()):
        dealers = state_data[state]
        wb = Workbook()
        ws = wb.active
        ws.title = state[:31].replace("/", "-").replace("\\", "-")

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = ha
            cell.border = tb
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + ci)].width = w

        for ri, d in enumerate(dealers, 2):
            ws.cell(row=ri, column=1, value=ri - 1).font = df
            ws.cell(row=ri, column=2, value=d.get("name", "")).font = df
            ws.cell(row=ri, column=3, value=d.get("phone", "")).font = df
            ws.cell(row=ri, column=4, value=d.get("address", "")).font = df
            ws.cell(row=ri, column=5, value=d.get("district", "")).font = df
            ws.cell(row=ri, column=6, value=d.get("city", "")).font = df
            ws.cell(row=ri, column=7, value=d.get("pincode", "")).font = df
            ws.cell(row=ri, column=8, value=d.get("state", "")).font = df
            ws.cell(row=ri, column=9, value=d.get("source", "")).font = df
            for c in range(1, 10):
                ws.cell(row=ri, column=c).border = tb

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:I{len(dealers) + 1}"

        safe = state.replace(" ", "_").replace("&", "and")
        fp = os.path.join(output_dir, f"{safe}_medical_dealers_JustDial.xlsx")
        wb.save(fp)
        print(f"  {state}: {len(dealers)} dealers -> {fp}")

    # Combined Excel with all states
    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet(title="Summary", index=0)
    sh = ["S.No", "State", "Districts", "Total Dealers"]
    for ci, h in enumerate(sh, 1):
        cell = summary.cell(row=1, column=ci, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = ha
        cell.border = tb
    summary.column_dimensions["A"].width = 8
    summary.column_dimensions["B"].width = 28
    summary.column_dimensions["C"].width = 12
    summary.column_dimensions["D"].width = 15

    row = 2
    total = 0
    for state in sorted(state_data.keys()):
        dealers = state_data[state]
        districts_count = len(set(d.get("district", "") for d in dealers))

        summary.cell(row=row, column=1, value=row - 1).font = df
        summary.cell(row=row, column=2, value=state).font = df
        summary.cell(row=row, column=3, value=districts_count).font = df
        summary.cell(row=row, column=4, value=len(dealers)).font = df
        for c in range(1, 5):
            summary.cell(row=row, column=c).border = tb

        sn = state[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(title=sn)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = ha
            cell.border = tb
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + ci)].width = w
        for ri, d in enumerate(dealers, 2):
            ws.cell(row=ri, column=1, value=ri - 1).font = df
            ws.cell(row=ri, column=2, value=d.get("name", "")).font = df
            ws.cell(row=ri, column=3, value=d.get("phone", "")).font = df
            ws.cell(row=ri, column=4, value=d.get("address", "")).font = df
            ws.cell(row=ri, column=5, value=d.get("district", "")).font = df
            ws.cell(row=ri, column=6, value=d.get("city", "")).font = df
            ws.cell(row=ri, column=7, value=d.get("pincode", "")).font = df
            ws.cell(row=ri, column=8, value=d.get("state", "")).font = df
            ws.cell(row=ri, column=9, value=d.get("source", "")).font = df
            for c in range(1, 10):
                ws.cell(row=ri, column=c).border = tb
        ws.freeze_panes = "A2"

        total += len(dealers)
        row += 1

    bf = Font(name="Calibri", bold=True, size=11)
    summary.cell(row=row, column=2, value="GRAND TOTAL").font = bf
    summary.cell(row=row, column=4, value=total).font = bf
    for c in range(1, 5):
        summary.cell(row=row, column=c).border = tb
        summary.cell(row=row, column=c).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    summary.freeze_panes = "A2"

    combined_path = os.path.join(output_dir, "JustDial_ALL_STATES_combined.xlsx")
    wb.save(combined_path)
    print(f"\n  Combined: {total} dealers -> {combined_path}")
    print("\nDone!")


if __name__ == "__main__":
    main()

"""
Medical Dealer Scraper - Google Maps (real phone numbers)
Headless Playwright, 2 queries per district, smart anti-detection.
Tamil Nadu first, then all states.
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import time
import random
import re
import json
import os
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from districts import STATES_DISTRICTS

PROGRESS_FILE = "scrape_progress.json"
OUTPUT_DIR = "output"

SEARCH_QUERIES = [
    "medical dealer in {district} {state}",
    "medicine distributor in {district} {state}",
]

MEDICAL_KEYWORDS = [
    "medical", "medicine", "medic", "pharma", "pharmacy", "drug",
    "pharmaceutical", "surgical", "healthcare", "health care",
    "hospital", "diagnostic", "lab", "laboratory", "dental", "ortho",
    "ayurved", "homeo", "herbal", "distributor", "dealer", "supplier",
    "wholesale", "equipment", "consumable", "supply", "instrument",
    "chemist", "dispensary", "surgical",
]


def is_medical(name, category=""):
    text = (name + " " + category).lower()
    return any(kw in text for kw in MEDICAL_KEYWORDS)


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"completed": [], "data": {}}


def save_progress(progress):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, indent=2, ensure_ascii=False)


def random_delay(a=3, b=6):
    time.sleep(random.uniform(a, b))


class BrowserManager:
    def __init__(self, pw):
        self.pw = pw
        self.browser = None
        self.page = None
        self.request_count = 0
        self._launch()

    def _launch(self):
        try:
            if self.browser:
                self.browser.close()
        except Exception:
            pass
        print("  [BROWSER] Launching...")
        self.browser = self.pw.chromium.launch(
            headless=False,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
            ],
        )
        ctx = self.browser.new_context(
            viewport={"width": 1366, "height": 900},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            locale="en-IN",
            timezone_id="Asia/Kolkata",
            geolocation={"latitude": 13.0827, "longitude": 80.2707},
            permissions=["geolocation"],
        )
        # Block heavy resources for speed
        ctx.route("**/*.{png,jpg,jpeg,gif,svg,woff,woff2,ttf}", lambda r: r.abort())
        self.page = ctx.new_page()
        # Anti-detection
        self.page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        """)
        self.request_count = 0
        print("  [BROWSER] Ready.")

    def goto(self, url, retries=3):
        for attempt in range(retries):
            try:
                self.page.evaluate("1+1")
            except Exception:
                self._launch()
                random_delay(3, 5)
            try:
                self.page.goto(url, timeout=60000, wait_until="domcontentloaded")
                self.request_count += 1
                # Restart browser every 50 requests to avoid detection
                if self.request_count > 50:
                    print("  [BROWSER] Rotating (50 requests)...")
                    self._launch()
                    random_delay(5, 10)
                    self.page.goto(url, timeout=60000, wait_until="domcontentloaded")
                return True
            except Exception as e:
                err = str(e)[:80]
                print(f"    goto fail ({attempt+1}): {err}")
                if "closed" in err.lower():
                    self._launch()
                random_delay(5, 10)
        return False

    def close(self):
        try:
            self.browser.close()
        except Exception:
            pass


def scroll_feed(page, scrolls=5):
    for _ in range(scrolls):
        try:
            page.evaluate('document.querySelector(\'div[role="feed"]\')?.scrollBy(0, 1000)')
            time.sleep(random.uniform(1.5, 2.5))
        except Exception:
            break


def scrape_google_maps(bm, district, state):
    """Scrape medical dealers from Google Maps for a district."""
    results = []
    seen_names = set()

    for query_tpl in SEARCH_QUERIES:
        query = query_tpl.format(district=district, state=state)
        url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"

        if not bm.goto(url):
            continue
        random_delay(4, 7)

        # Always get fresh page reference (may change after browser rotation)
        page = bm.page

        # Scroll to load more results
        scroll_feed(page, scrolls=5)

        # Get result cards
        cards = page.query_selector_all('a.hfpxzc')
        if not cards:
            cards = page.query_selector_all('div.Nv2PK')
        print(f"    [{len(cards)} results] {query[:50]}")

        for card in cards:
            try:
                # Click card to open detail panel
                card.click()
                random_delay(2, 4)

                # Name
                name_el = page.query_selector('h1.DUwDvf')
                if not name_el:
                    continue
                name = name_el.inner_text().strip()
                if not name or name.lower() in seen_names:
                    continue

                # Category
                category = ""
                cat_el = page.query_selector('button[jsaction="pane.rating.category"]')
                if cat_el:
                    category = cat_el.inner_text().strip()
                if not category:
                    cat_el2 = page.query_selector('span.DkEaL')
                    if cat_el2:
                        category = cat_el2.inner_text().strip()

                # Filter
                if not is_medical(name, category):
                    continue

                # PHONE - the key data we need
                phone = ""
                # Method 1: Copy phone button
                phone_btn = page.query_selector('button[data-tooltip="Copy phone number"]')
                if phone_btn:
                    aria = phone_btn.get_attribute("aria-label") or ""
                    phone = aria.replace("Phone:", "").replace("phone:", "").strip()

                # Method 2: data-item-id="phone" button
                if not phone:
                    phone_el = page.query_selector('button[data-item-id^="phone"]')
                    if phone_el:
                        aria = phone_el.get_attribute("aria-label") or ""
                        phone = aria.replace("Phone:", "").replace("phone:", "").strip()

                # Method 3: tel: link
                if not phone:
                    tel_link = page.query_selector('a[href^="tel:"]')
                    if tel_link:
                        phone = tel_link.get_attribute("href").replace("tel:", "").strip()

                # Method 4: Search info section for phone pattern
                if not phone:
                    info_els = page.query_selector_all('div.rogA2c div.Io6YTe')
                    for el in info_els:
                        txt = el.inner_text()
                        match = re.search(r'[\+]?[0-9][\s\-0-9]{8,14}', txt)
                        if match:
                            phone = match.group().strip()
                            break

                # Skip if no real phone
                if not phone:
                    continue

                # Clean phone - must look like a real number
                phone_digits = re.sub(r'\D', '', phone)
                if len(phone_digits) < 10:
                    continue

                # Address
                address = ""
                addr_btn = page.query_selector('button[data-item-id="address"]')
                if addr_btn:
                    address = (addr_btn.get_attribute("aria-label") or "").replace("Address:", "").strip()

                seen_names.add(name.lower())
                results.append({
                    "name": name,
                    "phone": phone,
                    "address": address,
                    "district": district,
                    "state": state,
                })

            except Exception as e:
                if "closed" in str(e).lower():
                    break
                continue

        random_delay(5, 10)  # Long delay between queries

    return results


def scrape_district(bm, state, district, progress):
    key = f"{state}|{district}"
    if key in progress["completed"]:
        return

    print(f"\n  -- {district}, {state} --")
    results = scrape_google_maps(bm, district, state)

    # Deduplicate
    seen = set()
    unique = []
    for r in results:
        k = r["name"].lower().strip()
        if k not in seen:
            seen.add(k)
            unique.append(r)

    progress["completed"].append(key)
    if state not in progress["data"]:
        progress["data"][state] = []
    progress["data"][state].extend(unique)
    save_progress(progress)

    with_phone = len([r for r in unique if r["phone"]])
    print(f"  => {len(unique)} dealers ({with_phone} with phone) in {district}")


# ──── EXCEL ────
def get_styles():
    return {
        "hf": Font(name="Calibri", bold=True, size=12, color="FFFFFF"),
        "fill": PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
        "align": Alignment(horizontal="center", vertical="center"),
        "border": Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin")),
        "df": Font(name="Calibri", size=11),
    }


def create_state_excel(state, dealers):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    s = get_styles()
    ws.title = state[:31].replace("/", "-")
    headers = ["S.No", "Name", "Phone", "Address", "District", "State"]
    widths = [8, 35, 20, 50, 22, 22]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = s["hf"], s["fill"], s["align"], s["border"]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + ci)].width = w
    for ri, d in enumerate(dealers, 2):
        ws.cell(row=ri, column=1, value=ri-1).font = s["df"]
        ws.cell(row=ri, column=2, value=d["name"]).font = s["df"]
        ws.cell(row=ri, column=3, value=d["phone"]).font = s["df"]
        ws.cell(row=ri, column=4, value=d.get("address", "")).font = s["df"]
        ws.cell(row=ri, column=5, value=d["district"]).font = s["df"]
        ws.cell(row=ri, column=6, value=d["state"]).font = s["df"]
        for c in range(1, 7):
            ws.cell(row=ri, column=c).border = s["border"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(dealers)+1}"

    safe = state.replace(" ", "_").replace("&", "and")
    fp = os.path.join(OUTPUT_DIR, f"{safe}_medical_dealers.xlsx")
    wb.save(fp)
    print(f"\n  >>> EXCEL: {fp} ({len(dealers)} dealers)")


def create_combined_excel(progress):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    s = get_styles()
    headers = ["S.No", "Name", "Phone", "Address", "District", "State"]
    widths = [8, 35, 20, 50, 22, 22]
    total = 0

    summary = wb.create_sheet(title="Summary", index=0)
    for ci, h in enumerate(["S.No", "State", "Dealers"], 1):
        c = summary.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = s["hf"], s["fill"], s["align"], s["border"]
    summary.column_dimensions["A"].width = 8
    summary.column_dimensions["B"].width = 28
    summary.column_dimensions["C"].width = 15

    row = 2
    for state in sorted(progress["data"].keys()):
        dealers = progress["data"][state]
        if not dealers:
            continue
        sn = state[:31].replace("/", "-")
        ws = wb.create_sheet(title=sn)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font, c.fill, c.alignment, c.border = s["hf"], s["fill"], s["align"], s["border"]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + ci)].width = w
        for ri, d in enumerate(dealers, 2):
            ws.cell(row=ri, column=1, value=ri-1).font = s["df"]
            ws.cell(row=ri, column=2, value=d["name"]).font = s["df"]
            ws.cell(row=ri, column=3, value=d["phone"]).font = s["df"]
            ws.cell(row=ri, column=4, value=d.get("address", "")).font = s["df"]
            ws.cell(row=ri, column=5, value=d["district"]).font = s["df"]
            ws.cell(row=ri, column=6, value=d["state"]).font = s["df"]
            for c in range(1, 7):
                ws.cell(row=ri, column=c).border = s["border"]
        ws.freeze_panes = "A2"

        summary.cell(row=row, column=1, value=row-1).font = s["df"]
        summary.cell(row=row, column=2, value=state).font = s["df"]
        summary.cell(row=row, column=3, value=len(dealers)).font = s["df"]
        for c in range(1, 4):
            summary.cell(row=row, column=c).border = s["border"]
        total += len(dealers)
        row += 1

    bf = Font(name="Calibri", bold=True, size=11)
    summary.cell(row=row, column=2, value="TOTAL").font = bf
    summary.cell(row=row, column=3, value=total).font = bf
    for c in range(1, 4):
        summary.cell(row=row, column=c).border = s["border"]
    summary.freeze_panes = "A2"

    fp = os.path.join(OUTPUT_DIR, "medical_dealers_all_india.xlsx")
    wb.save(fp)
    print(f"\n  Combined: {fp} ({total} dealers)")


def main():
    progress = load_progress()
    total_d = sum(len(d) for d in STATES_DISTRICTS.values())
    done = len(progress["completed"])

    print("=" * 60)
    print("  MEDICAL DEALER SCRAPER - Google Maps (Real Phone Numbers)")
    print(f"  States: {len(STATES_DISTRICTS)} | Districts: {total_d}")
    print(f"  Completed: {done} | Remaining: {total_d - done}")
    print(f"  Queries per district: {len(SEARCH_QUERIES)}")
    print(f"  Only dealers with REAL phone numbers are saved")
    print("=" * 60)

    ordered = {}
    if "Tamil Nadu" in STATES_DISTRICTS:
        ordered["Tamil Nadu"] = STATES_DISTRICTS["Tamil Nadu"]
    for s in sorted(STATES_DISTRICTS.keys()):
        if s != "Tamil Nadu":
            ordered[s] = STATES_DISTRICTS[s]

    with sync_playwright() as p:
        bm = BrowserManager(p)
        try:
            for state, districts in ordered.items():
                print(f"\n{'='*55}")
                print(f"STATE: {state} ({len(districts)} districts)")
                print(f"{'='*55}")

                for district in districts:
                    scrape_district(bm, state, district, progress)

                if state in progress["data"] and progress["data"][state]:
                    create_state_excel(state, progress["data"][state])

        except KeyboardInterrupt:
            print("\nInterrupted! Progress saved.")
            save_progress(progress)
        finally:
            bm.close()

    if progress["data"]:
        create_combined_excel(progress)
    print("\nDone!")


if __name__ == "__main__":
    main()

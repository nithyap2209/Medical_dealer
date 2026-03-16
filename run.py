"""
Medical Dealer Scraper - Runner
Scrapes JustDial + IndiaMART for medical dealers across all Indian districts.
Generates state-wise Excel files.

Usage:
    python run.py                           # All states (Tamil Nadu first)
    python run.py --state "Tamil Nadu"      # Single state only
    python run.py --spider justdial         # Only JustDial
    python run.py --spider indiamart        # Only IndiaMART
    python run.py --max-pages 3             # Limit pages per category
"""

import argparse
import json
import os
import sys
import subprocess


def get_districts():
    """Load districts from districts.py"""
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from districts import STATES_DISTRICTS
    return STATES_DISTRICTS


def order_states(states_districts, first_state="Tamil Nadu"):
    """Put Tamil Nadu first, then rest alphabetically."""
    ordered = {}
    if first_state in states_districts:
        ordered[first_state] = states_districts[first_state]
    for state in sorted(states_districts.keys()):
        if state != first_state:
            ordered[state] = states_districts[state]
    return ordered


def run_spider(spider_name, states_districts, max_pages=None):
    """Run a Scrapy spider for given states/districts."""
    # Write districts to temp file for spider to read
    temp_file = os.path.join("medical_scraper", "_districts_input.json")
    with open(temp_file, "w", encoding="utf-8") as f:
        json.dump(states_districts, f, ensure_ascii=False)

    cmd = [
        sys.executable, "-m", "scrapy", "crawl", spider_name,
        "-a", f"districts_file={os.path.abspath(temp_file)}",
    ]
    if max_pages:
        cmd.extend(["-a", f"max_pages={max_pages}"])

    print(f"\n  Running spider: {spider_name}")
    print(f"  States: {len(states_districts)}, Districts: {sum(len(d) for d in states_districts.values())}")

    result = subprocess.run(
        cmd,
        cwd=os.path.join(os.path.dirname(os.path.abspath(__file__)), "medical_scraper"),
    )
    return result.returncode


def generate_excel():
    """Generate state-wise Excel from scraped JSON data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data_file = os.path.join("output", "scraped_data.json")
    if not os.path.exists(data_file):
        print("No scraped data found!")
        return

    with open(data_file, "r", encoding="utf-8") as f:
        all_items = json.load(f)

    print(f"\nTotal records: {len(all_items)}")

    # Group by state
    state_data = {}
    for item in all_items:
        state = item.get("state", "Unknown")
        if state not in state_data:
            state_data[state] = []
        state_data[state].append(item)

    # Styles
    hf = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center")
    tb = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    df = Font(name="Calibri", size=11)

    headers = ["S.No", "Name", "Phone", "Address", "District", "City", "Pincode", "State", "Source"]
    col_widths = [8, 35, 18, 45, 20, 18, 10, 20, 12]

    os.makedirs("output", exist_ok=True)

    # Individual state Excel files
    for state in sorted(state_data.keys()):
        dealers = state_data[state]
        wb = Workbook()
        ws = wb.active
        ws.title = state[:31].replace("/", "-").replace("\\", "-")

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = ha
            cell.border = tb
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + ci)].width = w

        for ri, d in enumerate(dealers, 2):
            ws.cell(row=ri, column=1, value=ri - 1).font = df
            ws.cell(row=ri, column=2, value=d.get("name", "")).font = df
            ws.cell(row=ri, column=3, value=d.get("phone", "")).font = df
            ws.cell(row=ri, column=4, value=d.get("address", "")).font = df
            ws.cell(row=ri, column=5, value=d.get("district", "")).font = df
            ws.cell(row=ri, column=6, value=d.get("city", "")).font = df
            ws.cell(row=ri, column=7, value=d.get("pincode", "")).font = df
            ws.cell(row=ri, column=8, value=d.get("state", "")).font = df
            ws.cell(row=ri, column=9, value=d.get("source", "")).font = df
            for c in range(1, 10):
                ws.cell(row=ri, column=c).border = tb

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:I{len(dealers) + 1}"

        safe = state.replace(" ", "_").replace("&", "and")
        fp = os.path.join("output", f"{safe}_medical_dealers.xlsx")
        wb.save(fp)
        print(f"  {state}: {len(dealers)} dealers -> {fp}")

    # Combined Excel with all states as sheets
    wb = Workbook()
    wb.remove(wb.active)

    # Summary sheet
    summary = wb.create_sheet(title="Summary", index=0)
    sh = ["S.No", "State", "Districts", "Total Dealers"]
    for ci, h in enumerate(sh, 1):
        cell = summary.cell(row=1, column=ci, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = ha
        cell.border = tb
    summary.column_dimensions["A"].width = 8
    summary.column_dimensions["B"].width = 28
    summary.column_dimensions["C"].width = 12
    summary.column_dimensions["D"].width = 15

    row = 2
    total = 0
    for state in sorted(state_data.keys()):
        dealers = state_data[state]
        districts_count = len(set(d.get("district", "") for d in dealers))

        summary.cell(row=row, column=1, value=row - 1).font = df
        summary.cell(row=row, column=2, value=state).font = df
        summary.cell(row=row, column=3, value=districts_count).font = df
        summary.cell(row=row, column=4, value=len(dealers)).font = df
        for c in range(1, 5):
            summary.cell(row=row, column=c).border = tb

        # State data sheet
        sn = state[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(title=sn)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = ha
            cell.border = tb
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + ci)].width = w
        for ri, d in enumerate(dealers, 2):
            ws.cell(row=ri, column=1, value=ri - 1).font = df
            ws.cell(row=ri, column=2, value=d.get("name", "")).font = df
            ws.cell(row=ri, column=3, value=d.get("phone", "")).font = df
            ws.cell(row=ri, column=4, value=d.get("address", "")).font = df
            ws.cell(row=ri, column=5, value=d.get("district", "")).font = df
            ws.cell(row=ri, column=6, value=d.get("city", "")).font = df
            ws.cell(row=ri, column=7, value=d.get("pincode", "")).font = df
            ws.cell(row=ri, column=8, value=d.get("state", "")).font = df
            ws.cell(row=ri, column=9, value=d.get("source", "")).font = df
            for c in range(1, 10):
                ws.cell(row=ri, column=c).border = tb
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:I{len(dealers) + 1}"

        total += len(dealers)
        row += 1

    # Grand total
    bf = Font(name="Calibri", bold=True, size=11)
    summary.cell(row=row, column=2, value="GRAND TOTAL").font = bf
    summary.cell(row=row, column=4, value=total).font = bf
    for c in range(1, 5):
        summary.cell(row=row, column=c).border = tb
        summary.cell(row=row, column=c).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    summary.freeze_panes = "A2"

    combined_path = os.path.join("output", "medical_dealers_all_india.xlsx")
    wb.save(combined_path)
    print(f"\n  Combined: {total} dealers -> {combined_path}")


def main():
    parser = argparse.ArgumentParser(description="Medical Dealer Scraper")
    parser.add_argument("--state", type=str, help="Scrape only this state")
    parser.add_argument("--spider", type=str, choices=["justdial", "indiamart"], help="Run only this spider")
    parser.add_argument("--max-pages", type=int, default=5, help="Max pages per category")
    parser.add_argument("--excel-only", action="store_true", help="Only generate Excel from existing data")
    args = parser.parse_args()

    if args.excel_only:
        generate_excel()
        return

    all_districts = get_districts()

    if args.state:
        if args.state not in all_districts:
            print(f"State '{args.state}' not found!")
            print(f"Available: {', '.join(sorted(all_districts.keys()))}")
            return
        districts_to_scrape = {args.state: all_districts[args.state]}
    else:
        districts_to_scrape = order_states(all_districts)

    total_districts = sum(len(d) for d in districts_to_scrape.values())
    print("=" * 60)
    print("  MEDICAL DEALER SCRAPER - JustDial + IndiaMART")
    print(f"  States: {len(districts_to_scrape)} | Districts: {total_districts}")
    print("=" * 60)

    # Clear old data
    data_file = os.path.join("output", "scraped_data.json")
    if os.path.exists(data_file):
        os.remove(data_file)

    spiders = ["justdial", "indiamart"] if not args.spider else [args.spider]

    for spider in spiders:
        run_spider(spider, districts_to_scrape, args.max_pages)

    # Generate Excel
    print("\n" + "=" * 60)
    print("  GENERATING EXCEL FILES")
    print("=" * 60)
    generate_excel()
    print("\nDone!")


if __name__ == "__main__":
    main()
